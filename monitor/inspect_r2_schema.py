"""
inspect_r2_schema.py
====================
Reads every Excel file uploaded by the scraper pipeline from Cloudflare R2,
validates each file against the schema defined in websites-config.yml, and:

  1. Prints a pass/fail table for every scraper (today's date by default).
  2. Writes a JSON report back to R2  → {r2_prefix}/monitor/{partition-date}/report.json
     (partition date = listing date + 1 day, same as scraper Excel folders)
  3. Writes/updates monitor_stats.yml in the repo with observed real statistics
     (min/max row counts, actual file sizes, actual column sets, actual sheet names).
  4. Emits a GitHub Actions step-summary  ($GITHUB_STEP_SUMMARY).
  5. Fires webhook alerts when checks fail (MONITOR_ALERT_WEBHOOK_URL).

Usage
-----
  python monitor/inspect_r2_schema.py
  python monitor/inspect_r2_schema.py --date 2026-06-04
  python monitor/inspect_r2_schema.py --update-stats           # also write monitor_stats.yml
  python monitor/inspect_r2_schema.py --days-lookback 7        # sample last 7 days for stats
  python monitor/inspect_r2_schema.py --upload-config          # publish websites-config.yml to R2
  python monitor/inspect_r2_schema.py --upload-stats PATH      # bootstrap monitor_stats.yml in R2

Site identity is read from R2 (not the repo):
  monitor-sites/{MONITOR_SITE_SLUG}/site.yml

Required env vars
-----------------
  CF_R2_ACCESS_KEY_ID, CF_R2_SECRET_ACCESS_KEY,
  CF_R2_ENDPOINT_URL,  CF_R2_BUCKET_NAME
  MONITOR_SITE_SLUG    — folder name under monitor-sites/ (e.g. 4sale, boshamlan)

Optional env vars
-----------------
  MONITOR_ALERT_WEBHOOK_URL  — n8n / Slack / Discord / Telegram webhook for failure alerts
"""

import argparse
import boto3
import io
import json
import logging
import os
import re
import sys
import tempfile
import urllib.error
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import yaml

from monitor_r2 import (
    build_r2_client,
    load_site_config_from_r2,
    monitor_data_keys,
    partition_date_for_listing,
    report_r2_key,
    resolve_site_folder,
)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("monitor")

# ── Paths ─────────────────────────────────────────────────────────────────────
REPO_ROOT   = Path(__file__).resolve().parent.parent
CONFIG_FILE = REPO_ROOT / "websites-config.yml"


def list_excel_files(client, bucket: str, prefix: str) -> List[Dict]:
    """
    Return all .xlsx objects under *prefix*.
    Each item: {key, size, last_modified}
    Handles pagination automatically.
    """
    results = []
    paginator = client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            if obj["Key"].endswith(".xlsx"):
                results.append(
                    {
                        "key":           obj["Key"],
                        "size_bytes":    obj["Size"],
                        "last_modified": obj["LastModified"].isoformat(),
                    }
                )
    return results


def download_excel(client, bucket: str, key: str) -> Optional[bytes]:
    """Download an .xlsx file from R2 and return raw bytes."""
    try:
        resp = client.get_object(Bucket=bucket, Key=key)
        return resp["Body"].read()
    except Exception as exc:
        log.warning(f"Could not download {key}: {exc}")
        return None


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL INSPECTION
# ══════════════════════════════════════════════════════════════════════════════

def inspect_excel(raw: bytes, file_key: str) -> Dict:
    """
    Open an xlsx from bytes and return:
      sheets: [{name, row_count, columns: [str]}]
      readable: bool
      error: str | None
    """
    result = {"file_key": file_key, "readable": False, "sheets": [], "error": None}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Read header row
            headers = []
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row:
                headers = [str(c) for c in header_row if c is not None]
            # Count data rows (fast — don't load everything into memory)
            row_count = sum(1 for _ in rows_iter)
            result["sheets"].append(
                {"name": sheet_name, "row_count": row_count, "columns": headers}
            )
        wb.close()
        result["readable"] = True
    except Exception as exc:
        result["error"] = str(exc)
        log.warning(f"Error inspecting {file_key}: {exc}")
    return result


def check_data_quality(raw: bytes, sheet_name: str, listing_date: Optional[str] = None) -> Dict:
    """
    Run deeper data-quality checks on a single sheet using pandas.
    Returns a dict of quality metrics.
    """
    metrics = {
        "duplicate_id_count": 0,
        "null_id_pct": 0.0,
        "null_title_pct": 0.0,
        "null_price_pct": 0.0,
        "stale_date_pct": 0.0,   # rows where date_published is NOT yesterday
        "total_rows": 0,
    }
    try:
        df = pd.read_excel(io.BytesIO(raw), sheet_name=sheet_name, engine="openpyxl")
        metrics["total_rows"] = len(df)
        if df.empty:
            return metrics

        # Duplicate IDs
        id_col = next((c for c in df.columns if str(c).lower() in ("id", "listing id")), None)
        if id_col:
            metrics["null_id_pct"]      = round(df[id_col].isna().mean() * 100, 1)
            metrics["duplicate_id_count"] = int(df[id_col].duplicated().sum())

        # Null title
        title_col = next((c for c in df.columns if str(c).lower() == "title"), None)
        if title_col:
            metrics["null_title_pct"] = round(df[title_col].isna().mean() * 100, 1)

        # Null price
        price_col = next((c for c in df.columns if str(c).lower() == "price"), None)
        if price_col:
            metrics["null_price_pct"] = round(df[price_col].isna().mean() * 100, 1)

        # Date freshness — date_published should be yesterday
        date_col = next(
            (c for c in df.columns if str(c).lower() in ("date_published", "date published")),
            None,
        )
        if date_col:
            expected = listing_date or (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
            stale = df[date_col].astype(str).apply(
                lambda v: not str(v).startswith(expected)
            )
            metrics["stale_date_pct"] = round(stale.mean() * 100, 1)

    except Exception as exc:
        log.debug(f"Quality check skipped for sheet '{sheet_name}': {exc}")

    return metrics


# ══════════════════════════════════════════════════════════════════════════════
# CONFIG LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load_config(client=None, bucket: Optional[str] = None, config_key: Optional[str] = None) -> Dict:
    """Load config from the local file (dev) or R2 (CI when the file is gitignored)."""
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, encoding="utf-8") as fh:
            log.info(f"Loaded config from {CONFIG_FILE}")
            return yaml.safe_load(fh)

    if client and bucket and config_key:
        try:
            resp = client.get_object(Bucket=bucket, Key=config_key)
            raw  = resp["Body"].read().decode("utf-8")
            log.info(f"Loaded config from r2://{bucket}/{config_key}")
            return yaml.safe_load(raw)
        except client.exceptions.NoSuchKey:
            log.debug(f"No config at r2://{bucket}/{config_key}")
        except Exception as exc:
            log.warning(f"Could not load config from R2: {exc}")

    key_hint = config_key or "{r2_prefix}/monitor/websites-config.yml"
    raise FileNotFoundError(
        f"{CONFIG_FILE.name} not found locally and not available in R2 "
        f"at {key_hint}. Create the file locally, then run with "
        f"--upload-config to publish it to R2 for CI."
    )


def upload_config(client, bucket: str, config_key: str) -> None:
    """Upload the local websites-config.yml to R2 for GitHub Actions."""
    if not CONFIG_FILE.exists():
        raise FileNotFoundError(f"Cannot upload — {CONFIG_FILE} does not exist.")

    body = CONFIG_FILE.read_bytes()
    client.put_object(
        Bucket=bucket,
        Key=config_key,
        Body=body,
        ContentType="text/yaml",
    )
    log.info(f"Config uploaded → r2://{bucket}/{config_key}")


def r2_base_prefix(r2_path_raw: str) -> str:
    """
    Convert config r2_path like '{r2_bucket}/4sale-data/animals'
    to the actual in-bucket prefix  '4sale-data/animals'.
    """
    # Remove leading placeholder segment  e.g.  '{r2_bucket}/'
    path = r2_path_raw.strip()
    if path.startswith("{"):
        path = path.split("/", 1)[1] if "/" in path else path
    return path.strip("/")


def partition_date_for_data_date(dt: datetime) -> datetime:
    """R2 folder uses save_date = listing date + 1 day (yesterday's listings → today's partition)."""
    return partition_date_for_listing(dt)


def excel_prefixes_for_date(base: str, dt: datetime) -> List[str]:
    """
    Build R2 date-partition prefixes for Excel discovery.
    Tries zero-padded (month=06/day=09) and unpadded (month=6/day=9) forms.
    """
    seen: set = set()
    prefixes: List[str] = []
    for month in (f"{dt.month:02d}", str(dt.month)):
        for day in (f"{dt.day:02d}", str(dt.day)):
            prefix = f"{base}/year={dt.year}/month={month}/day={day}/"
            if prefix not in seen:
                seen.add(prefix)
                prefixes.append(prefix)
    return prefixes


# ══════════════════════════════════════════════════════════════════════════════
# VALIDATION AGAINST SCHEMA
# ══════════════════════════════════════════════════════════════════════════════

# Sentinel patterns used as sheet-name templates in the YAML
_TEMPLATE_MARKERS = ("{", "or Main", "or All Listings")

# Scrapers that may legitimately produce zero files on many days (e.g. no yesterday listings)
_FILES_OPTIONAL_SCRAPERS = frozenset({"New Car"})

# Per-scraper validation overrides when R2 schema does not match actual Excel layout
_SCRAPER_PROFILES: Dict[str, Dict] = {
    "Commercials": {
        "skip_info_sheet": True,
        "data_sheet_aliases": {"Sheet1", "Main", "Data", "Listings"},
        "core_columns": ["id", "title"],
    },
}

# Normalized column names that are enrichment / metadata — missing is OK
_OPTIONAL_CANONICAL_COLUMNS = frozenset({
    "images", "imagespaths", "s3images", "r2images", "imagescount",
    "imageurls",
    "specificationen", "specificationar",
    "slug", "daterelative", "datecreated", "dateexpired",
    "saveddate", "scrapeddate",
    "categoryarabic", "categoryenglish", "hassubcategories", "subcategoriescount",
    "userphone", "useremail", "fulladdressen", "fulladdress",
    "membership", "isverified", "description",
    "views", "latitude", "longitude",
})

# Map normalized header variants → canonical name for fuzzy matching
_COLUMN_CANONICAL: Dict[str, str] = {
    "listingid": "id",
    "savedtos3date": "saveddate",
    "savedtor2date": "saveddate",
    "datascrapeddate": "scrapeddate",
    "s3images": "images",
    "r2images": "images",
    "s3imagespaths": "imagespaths",
    "r2imagespaths": "imagespaths",
    "displaytitle": "name",
    "displaydescription": "about",
    "username": "username",
    "usertype": "usertype",
    "datepublished": "datepublished",
    "relativedate": "daterelative",
    "daterelative": "daterelative",
    "userads": "userads",
    "imagescount": "imagescount",
    "userid": "userid",
    "user": "user",
    "district": "district",
    "imageurls": "imageurls",
    "datecreated": "datecreated",
    "dateexpired": "dateexpired",
    "datesort": "datesort",
    "datepublished": "datepublished",
    "pmenabled": "pmenabled",
    "viewsno": "views",
    "specificationen": "specificationen",
    "specificationar": "specificationar",
    "fulladdressen": "fulladdress",
    "descriptionen": "description",
    "descriptionar": "description",
    "phonenumber": "phone",
    "whatsappphone": "phone",
}


def normalize_column_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def canonical_column(name: str) -> str:
    norm = normalize_column_name(name)
    return _COLUMN_CANONICAL.get(norm, norm)


def build_observed_column_index(columns: List[str]) -> set:
    return {canonical_column(c) for c in columns}


def resolve_required_columns(
    required: List[str],
    observed_columns: List[str],
    schema_optional: Optional[List[str]] = None,
) -> Tuple[List[str], List[str]]:
    """
    Return (missing_core, missing_optional) after fuzzy + alias matching.
    """
    observed = build_observed_column_index(observed_columns)
    schema_opt = {canonical_column(c) for c in (schema_optional or [])}
    missing_core: List[str] = []
    missing_optional: List[str] = []

    for col in required:
        canon = canonical_column(col)
        if canon in observed:
            continue
        if canon in _OPTIONAL_CANONICAL_COLUMNS or canon in schema_opt:
            missing_optional.append(col)
        else:
            missing_core.append(col)

    return missing_core, missing_optional


def is_template(name: str) -> bool:
    return any(m in name for m in _TEMPLATE_MARKERS)


def validate_file(
    inspected: Dict,
    schema_entry: Dict,
    scraper_name: str = "",
) -> Dict:
    """
    Compare an inspected file against its schema entry.
    Returns a validation result dict:
      passed: bool
      checks: [{name, passed, detail}]
    """
    checks = []
    profile = _SCRAPER_PROFILES.get(scraper_name, {})

    def add(name, passed, detail=""):
        checks.append({"name": name, "passed": passed, "detail": detail})

    # --- Readability ---
    add("file_readable", inspected["readable"], inspected.get("error", ""))

    if not inspected["readable"]:
        return {"passed": False, "checks": checks}

    observed_sheets = {s["name"]: s for s in inspected["sheets"]}
    schema_sheets   = schema_entry.get("sheets", [])

    def validate_sheet_columns(sheet_label: str, obs: Dict, req_cols: List[str], schema_sheet: Dict):
        if obs["row_count"] == 0 and not obs["columns"]:
            add(f"rows_in_{sheet_label[:20]}", True, "empty sheet — skipped")
            return
        if obs["row_count"] == 0:
            add(f"rows_in_{sheet_label[:20]}", True, "0 rows — column check skipped")
            return

        schema_optional = schema_sheet.get("optional_columns", [])
        missing_core, missing_optional = resolve_required_columns(
            req_cols, obs["columns"], schema_optional
        )
        add(
            f"columns_in_{sheet_label[:20]}",
            len(missing_core) == 0,
            (
                f"Missing core: {missing_core}"
                + (f"; optional absent: {missing_optional}" if missing_optional else "")
                + (
                    f"; actual headers: {obs['columns'][:12]}"
                    + ("…" if len(obs["columns"]) > 12 else "")
                    if missing_core else ""
                )
            ) if missing_core or missing_optional else "",
        )

    for schema_sheet in schema_sheets:
        sname    = schema_sheet["name"]
        req_cols = schema_sheet.get("required_columns", [])
        row_min, row_max = schema_sheet.get("row_count_range", [0, 999999])

        if profile.get("skip_info_sheet") and sname == "Info":
            continue

        # Template sheets  (e.g. "{child_category or Main}") — validate
        # all non-Info observed sheets against this template
        if is_template(sname):
            data_sheets = [
                s for s in inspected["sheets"]
                if s["name"] not in ("Info", "No Data")
            ]
            if profile.get("data_sheet_aliases"):
                data_sheets = [
                    s for s in inspected["sheets"]
                    if s["name"] in profile["data_sheet_aliases"]
                    or (s["name"] not in ("Info", "No Data") and s["row_count"] > 0)
                ]
            if not data_sheets:
                add("data_sheets_exist", False, "No data sheets found")
                continue
            core_cols = profile.get("core_columns") or req_cols
            for ds in data_sheets:
                if ds["row_count"] == 0:
                    add(f"rows_in_{ds['name'][:20]}", True, "0 rows — skipped")
                    continue
                validate_sheet_columns(ds["name"], ds, core_cols, schema_sheet)
                in_range = row_min <= ds["row_count"] <= row_max
                add(
                    f"rows_in_{ds['name'][:20]}",
                    in_range,
                    f"{ds['row_count']} rows (expected {row_min}–{row_max})",
                )
        else:
            # Exact sheet name — allow profile aliases (e.g. Commercials → Sheet1)
            obs = observed_sheets.get(sname)
            if obs is None and profile.get("data_sheet_aliases"):
                for alias in profile["data_sheet_aliases"]:
                    if alias in observed_sheets:
                        obs = observed_sheets[alias]
                        break
            if obs is None:
                add(f"sheet_exists_{sname}", False, f"Sheet '{sname}' not found")
                continue
            core_cols = profile.get("core_columns") if profile.get("core_columns") and sname != "Info" else req_cols
            validate_sheet_columns(sname, obs, core_cols, schema_sheet)
            in_range = row_min <= obs["row_count"] <= row_max
            add(
                f"rows_in_{sname[:20]}",
                in_range,
                f"{obs['row_count']} rows (expected {row_min}–{row_max})",
            )

    # --- File size ---
    min_kb  = schema_entry.get("min_file_size_kb", 0)
    size_kb = inspected.get("size_bytes", 0) / 1024
    add(
        "file_size_ok",
        size_kb >= min_kb,
        f"{size_kb:.1f} KB (min {min_kb} KB)",
    )

    passed = all(c["passed"] for c in checks)
    return {"passed": passed, "checks": checks}


def validate_observed_floor(
    inspected: Dict,
    scraper_name: str,
    alert_cfg: Dict,
) -> Dict:
    """Compare file size against per-scraper floors derived from real monitor_stats."""
    checks = []
    floors = alert_cfg.get("observed_min_file_size_kb", {})
    floor_kb = floors.get(scraper_name)
    if floor_kb is None:
        return {"passed": True, "checks": checks}

    size_kb = inspected.get("size_bytes", 0) / 1024
    passed = size_kb >= floor_kb
    checks.append({
        "name": "observed_file_size_floor",
        "passed": passed,
        "detail": f"{size_kb:.1f} KB (observed floor {floor_kb} KB)",
    })
    return {"passed": passed, "checks": checks}


def validate_trends(
    inspected: Dict,
    scraper_name: str,
    stats_entry: Optional[Dict],
    trend_cfg: Dict,
) -> Dict:
    """
    Compare today's file against historical monitor_stats.yml baselines.
    Skipped until min_observed_days of history exist for the scraper.
    """
    checks = []
    if not stats_entry:
        return {"passed": True, "checks": checks}

    min_days = trend_cfg.get("min_observed_days", 3)
    if len(stats_entry.get("observed_dates", [])) < min_days:
        return {"passed": True, "checks": checks}

    row_min_pct = trend_cfg.get("min_row_pct_of_hist_min", 30) / 100.0
    row_max_pct = trend_cfg.get("max_row_pct_of_hist_max", 300) / 100.0
    size_min_pct = trend_cfg.get("min_file_size_pct_of_hist_min", 50) / 100.0

    size_kb = inspected.get("size_bytes", 0) / 1024
    hist_size_min = stats_entry.get("file_size_kb", {}).get("min")
    if hist_size_min is not None:
        threshold = hist_size_min * size_min_pct
        passed = size_kb >= threshold
        checks.append({
            "name": "trend_file_size",
            "passed": passed,
            "detail": f"{size_kb:.1f} KB vs {threshold:.1f} KB ({size_min_pct:.0%} of hist min {hist_size_min} KB)",
        })

    sheet_stats = stats_entry.get("sheets", {})
    for sheet in inspected.get("sheets", []):
        sname = sheet["name"]
        if sname == "Info":
            continue
        rc = sheet["row_count"]
        sh = sheet_stats.get(sname, {}).get("row_count", {})
        hist_min = sh.get("min")
        hist_max = sh.get("max")

        if hist_min is not None and hist_min > 0:
            floor = max(1, int(hist_min * row_min_pct))
            passed = rc >= floor
            checks.append({
                "name": f"trend_rows_low_{sname[:18]}",
                "passed": passed,
                "detail": f"{rc} rows vs floor {floor} ({row_min_pct:.0%} of hist min {hist_min})",
            })

        if hist_max is not None and hist_max > 0:
            ceiling = int(hist_max * row_max_pct)
            passed = rc <= ceiling
            checks.append({
                "name": f"trend_rows_high_{sname[:18]}",
                "passed": passed,
                "detail": f"{rc} rows vs ceiling {ceiling} ({row_max_pct:.0%} of hist max {hist_max})",
            })

    passed = all(c["passed"] for c in checks)
    return {"passed": passed, "checks": checks}


def merge_validations(*parts: Dict) -> Dict:
    """Merge multiple validation result dicts into one."""
    checks: List[Dict] = []
    for part in parts:
        checks.extend(part.get("checks", []))
    return {"passed": all(c["passed"] for c in checks), "checks": checks}


def severity_for_check(check_name: str) -> str:
    if check_name in ("file_readable",) or check_name.startswith("sheet_exists"):
        return "critical"
    if check_name.startswith("trend_rows_low") or check_name == "trend_file_size":
        return "high"
    if check_name.startswith("trend_rows_high"):
        return "medium"
    if check_name == "observed_file_size_floor":
        return "high"
    return "high"


def collect_alerts(all_results: List[Dict], run_date: str) -> List[Dict]:
    """Build a flat list of alert events from scraper results."""
    alerts: List[Dict] = []
    for r in all_results:
        scraper = r["scraper"]
        if r["files_found"] == 0:
            if r.get("files_optional"):
                continue
            alerts.append({
                "scraper": scraper,
                "severity": "critical",
                "type": "no_files",
                "check": "files_found",
                "detail": "No Excel files found in R2 for target date",
                "file": None,
            })
            continue

        for fr in r.get("file_results", []):
            file_key = fr.get("file_key") or fr.get("validation", {}).get("file")
            for check in fr.get("validation", {}).get("checks", []):
                if check["passed"]:
                    continue
                alerts.append({
                    "scraper": scraper,
                    "severity": severity_for_check(check["name"]),
                    "type": "validation",
                    "check": check["name"],
                    "detail": check.get("detail", ""),
                    "file": file_key,
                })

            if fr.get("readable") and fr.get("sheets"):
                for sheet in fr["sheets"]:
                    q = sheet.get("quality")
                    if not q:
                        continue
                    if q.get("null_id_pct", 0) >= 5:
                        alerts.append({
                            "scraper": scraper,
                            "severity": "high",
                            "type": "quality",
                            "check": "null_id_pct",
                            "detail": f"Sheet '{sheet['name']}': {q['null_id_pct']}% null IDs",
                            "file": file_key,
                        })
                    if q.get("stale_date_pct", 0) >= 20:
                        alerts.append({
                            "scraper": scraper,
                            "severity": "high",
                            "type": "quality",
                            "check": "stale_date_pct",
                            "detail": f"Sheet '{sheet['name']}': {q['stale_date_pct']}% stale dates",
                            "file": file_key,
                        })
    return alerts


def should_send_alerts(alerts: List[Dict], alert_cfg: Dict, total_scrapers: int) -> bool:
    if not alert_cfg.get("enabled", True):
        return False
    if not alerts:
        return False
    if alert_cfg.get("alert_on_any_failure", True):
        max_rate = alert_cfg.get("max_scraper_fail_rate_pct", 0)
        if max_rate <= 0:
            return True
        failed_scrapers = len({a["scraper"] for a in alerts})
        rate = (failed_scrapers / total_scrapers * 100) if total_scrapers else 100
        return rate >= max_rate
    return False


def format_alert_text(run_date: str, alerts: List[Dict], passed: int, total: int) -> str:
    lines = [
        f"4Sale Schema Monitor ALERT — {run_date}",
        f"{passed}/{total} scrapers passed · {len(alerts)} issue(s)",
        "",
    ]
    by_scraper: Dict[str, List[Dict]] = defaultdict(list)
    for a in alerts:
        by_scraper[a["scraper"]].append(a)
    for scraper, items in sorted(by_scraper.items()):
        lines.append(f"[{scraper}]")
        for a in items[:8]:
            file_bit = f" — {a['file'].split('/')[-1]}" if a.get("file") else ""
            lines.append(f"  • {a['severity'].upper()} {a['check']}: {a['detail']}{file_bit}")
        if len(items) > 8:
            lines.append(f"  … +{len(items) - 8} more")
        lines.append("")
    return "\n".join(lines).strip()


def format_alert_html(run_date: str, alerts: List[Dict], passed: int, total: int) -> str:
    """HTML body for n8n email node."""
    rows = []
    for a in alerts[:50]:
        file_cell = a.get("file", "") or "—"
        if file_cell != "—":
            file_cell = file_cell.split("/")[-1]
        rows.append(
            f"<tr>"
            f"<td>{a['scraper']}</td>"
            f"<td>{a['severity']}</td>"
            f"<td>{a['check']}</td>"
            f"<td>{a.get('detail', '')}</td>"
            f"<td>{file_cell}</td>"
            f"</tr>"
        )
    extra = ""
    if len(alerts) > 50:
        extra = f"<p><em>… and {len(alerts) - 50} more issues</em></p>"

    failed = sorted({a["scraper"] for a in alerts})
    return (
        f"<h2>4Sale Schema Monitor ALERT — {run_date}</h2>"
        f"<p><strong>{passed}/{total}</strong> scrapers passed · "
        f"<strong>{len(alerts)}</strong> issue(s)</p>"
        f"<p>Failed scrapers: {', '.join(failed)}</p>"
        f"<table border='1' cellpadding='6' cellspacing='0' "
        f"style='border-collapse:collapse;font-family:sans-serif;font-size:13px'>"
        f"<tr style='background:#f0f0f0'>"
        f"<th>Scraper</th><th>Severity</th><th>Check</th><th>Detail</th><th>File</th>"
        f"</tr>"
        + "".join(rows)
        + f"</table>{extra}"
    )


def send_webhook_alert(payload: Dict, url: str) -> None:
    """POST JSON alert payload to webhook URL (n8n / Slack / Discord / Telegram)."""
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            log.info(f"Alert webhook sent — HTTP {resp.status}")
    except urllib.error.HTTPError as exc:
        log.error(f"Alert webhook HTTP error {exc.code}: {exc.read().decode()[:500]}")
        raise
    except Exception as exc:
        log.error(f"Alert webhook failed: {exc}")
        raise


def dispatch_alerts(
    alerts: List[Dict],
    all_results: List[Dict],
    run_date: str,
    alert_cfg: Dict,
    webhook_url: Optional[str],
    site: Dict,
    meta: Dict,
) -> None:
    if not webhook_url:
        log.info("MONITOR_ALERT_WEBHOOK_URL not set — alerts logged only.")
        return

    channels = alert_cfg.get("channels", {}).get("webhook", {})
    if not channels.get("enabled", True):
        log.info("Webhook alerts disabled in alert_thresholds.channels.webhook.")
        return

    total = len(all_results)
    passed = sum(1 for r in all_results if r["all_passed"])
    if not should_send_alerts(alerts, alert_cfg, total):
        log.info("Alert threshold not met — webhook skipped.")
        return

    display = site.get("display_name") or meta.get("website") or site.get("site_id", "Monitor")
    website = meta.get("website") or site.get("website", "")

    payload = {
        "source": f"{site.get('site_id', 'site')}-schema-monitor",
        "site_id": site.get("site_id"),
        "website": website,
        "country": meta.get("country") or site.get("country"),
        "repo": meta.get("repo") or site.get("repo"),
        "display_name": display,
        "status": "failed",
        "run_date": run_date,
        "summary": f"{passed}/{total} scrapers passed",
        "alert_count": len(alerts),
        "text": format_alert_text(run_date, alerts, passed, total),
        "email_subject": f"[{display}] ALERT {run_date} — {passed}/{total} passed, {len(alerts)} issue(s)",
        "email_html": format_alert_html(run_date, alerts, passed, total),
        "alerts": alerts,
        "scrapers_failed": sorted({a["scraper"] for a in alerts}),
    }
    send_webhook_alert(payload, webhook_url)


def upload_stats_file(client, bucket: str, local_path: Path, stats_key: str) -> None:
    """Bootstrap monitor_stats.yml in R2 from a local file."""
    body = local_path.read_bytes()
    client.put_object(
        Bucket=bucket,
        Key=stats_key,
        Body=body,
        ContentType="text/yaml",
    )
    log.info(f"Stats uploaded → r2://{bucket}/{stats_key}")


# ══════════════════════════════════════════════════════════════════════════════
# STATS ACCUMULATION & YAML WRITER
# ══════════════════════════════════════════════════════════════════════════════

def accumulate_stats(
    existing: Dict,
    scraper_name: str,
    inspected: Dict,
    size_bytes: int,
    run_date: str,
) -> None:
    """
    Merge new observations into the *existing* stats dict (in-place).
    Structure:
      scraper_name:
        observed_dates: [...]
        file_size_kb: {min, max, last}
        sheets:
          sheet_name:
            row_count: {min, max, last}
            columns: [...]
    """
    entry = existing.setdefault(scraper_name, {
        "observed_dates":  [],
        "file_size_kb":    {"min": None, "max": None, "last": None},
        "sheets":          {},
    })

    if run_date not in entry["observed_dates"]:
        entry["observed_dates"].append(run_date)
        entry["observed_dates"] = sorted(entry["observed_dates"])[-30:]  # keep last 30

    kb = round(size_bytes / 1024, 1)
    fs = entry["file_size_kb"]
    fs["last"] = kb
    fs["min"]  = min(kb, fs["min"]) if fs["min"] is not None else kb
    fs["max"]  = max(kb, fs["max"]) if fs["max"] is not None else kb

    for sheet in inspected.get("sheets", []):
        sname = sheet["name"]
        se    = entry["sheets"].setdefault(sname, {
            "row_count": {"min": None, "max": None, "last": None},
            "columns":   [],
        })
        rc = sheet["row_count"]
        se["row_count"]["last"] = rc
        se["row_count"]["min"]  = min(rc, se["row_count"]["min"]) if se["row_count"]["min"] is not None else rc
        se["row_count"]["max"]  = max(rc, se["row_count"]["max"]) if se["row_count"]["max"] is not None else rc
        # Merge new columns (union)
        new_cols = [c for c in sheet["columns"] if c not in se["columns"]]
        se["columns"].extend(new_cols)


def load_existing_stats(client, bucket: str, stats_key: str) -> Dict:
    """Download monitor_stats.yml from R2 and parse it. Returns {} if not found."""
    try:
        resp = client.get_object(Bucket=bucket, Key=stats_key)
        raw  = resp["Body"].read().decode("utf-8")
        data = yaml.safe_load(raw) or {}
        log.info(f"Loaded existing stats from r2://{bucket}/{stats_key}")
        return data
    except client.exceptions.NoSuchKey:
        log.info(f"No existing monitor_stats.yml at r2://{bucket}/{stats_key}")
        return {}
    except Exception as exc:
        log.warning(f"Could not load stats from R2: {exc} — starting fresh.")
        return {}


def save_stats(client, bucket: str, stats: Dict, stats_key: str) -> None:
    """Serialise stats to YAML and upload to R2 (overwrites previous version)."""
    header = (
        "# Auto-generated by monitor/inspect_r2_schema.py\n"
        "# Stored in Cloudflare R2 — do not add to the repo.\n"
        "# Re-run the weekly monitor workflow to refresh.\n\n"
    )
    body = (header + yaml.dump(
        stats, allow_unicode=True, default_flow_style=False, sort_keys=True
    )).encode("utf-8")
    try:
        client.put_object(
            Bucket=bucket,
            Key=stats_key,
            Body=body,
            ContentType="text/yaml",
        )
        log.info(f"Saved observed stats → r2://{bucket}/{stats_key}")
    except Exception as exc:
        log.error(f"Failed to save stats to R2: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# REPORTING
# ══════════════════════════════════════════════════════════════════════════════

PASS = "✅"
FAIL = "❌"
WARN = "⚠️"
MISS = "🚫"

# Max failed-check lines logged per scraper (use --verbose for all)
_DEFAULT_MAX_FAILURE_LOGS = 25


def collect_file_failures(scraper_result: Dict) -> List[Dict]:
    """Flatten failed validation + quality checks for one scraper."""
    failures: List[Dict] = []
    for fr in scraper_result.get("file_results", []):
        file_key = fr.get("file_key") or fr.get("validation", {}).get("file", "?")
        file_name = file_key.split("/")[-1] if file_key else "?"

        for check in fr.get("validation", {}).get("checks", []):
            if not check["passed"]:
                failures.append({
                    "file": file_name,
                    "file_key": file_key,
                    "check": check["name"],
                    "detail": check.get("detail", ""),
                    "kind": "validation",
                })

        if fr.get("readable") and fr.get("sheets"):
            for sheet in fr["sheets"]:
                q = sheet.get("quality")
                if not q:
                    continue
                if q.get("null_id_pct", 0) >= 5:
                    failures.append({
                        "file": file_name,
                        "file_key": file_key,
                        "check": "null_id_pct",
                        "detail": f"Sheet '{sheet['name']}': {q['null_id_pct']}% null IDs",
                        "kind": "quality",
                    })
                if q.get("stale_date_pct", 0) >= 20:
                    failures.append({
                        "file": file_name,
                        "file_key": file_key,
                        "check": "stale_date_pct",
                        "detail": f"Sheet '{sheet['name']}': {q['stale_date_pct']}% stale dates",
                        "kind": "quality",
                    })
    return failures


def log_scraper_failures(
    scraper_name: str,
    scraper_result: Dict,
    max_lines: Optional[int] = _DEFAULT_MAX_FAILURE_LOGS,
) -> None:
    """Log which checks failed and why (per file)."""
    if scraper_result["files_found"] == 0:
        return

    failures = collect_file_failures(scraper_result)
    if not failures:
        return

    passed = scraper_result["checks_passed"]
    total = scraper_result["checks_total"]
    log.warning(
        f"  {scraper_name} — {len(failures)} failed check(s) "
        f"({passed}/{total} passed):"
    )
    show = failures if max_lines is None else failures[:max_lines]
    for item in show:
        detail = item["detail"]
        detail_bit = f": {detail}" if detail else ""
        log.warning(f"    • {item['file']} → {item['check']}{detail_bit}")
    if max_lines is not None and len(failures) > max_lines:
        log.warning(f"    … +{len(failures) - max_lines} more failure(s) (use --verbose for all)")


def print_failure_summary(results: List[Dict]) -> None:
    """Print a cross-scraper breakdown of the most common failure types."""
    by_check: Dict[str, List[str]] = defaultdict(list)
    no_files: List[str] = []

    for r in results:
        if r["files_found"] == 0:
            no_files.append(r["scraper"])
            continue
        for item in collect_file_failures(r):
            by_check[item["check"]].append(r["scraper"])

    if not no_files and not by_check:
        return

    print("\n" + "-" * 70)
    print("FAILURE BREAKDOWN")
    print("-" * 70)

    if no_files:
        print(f"\nNo Excel files ({len(no_files)}):")
        for name in no_files:
            print(f"  • {name}")

    if by_check:
        print(f"\nFailed checks by type ({sum(len(v) for v in by_check.values())} total):")
        ranked = sorted(by_check.items(), key=lambda kv: len(kv[1]), reverse=True)
        for check_name, scrapers in ranked[:20]:
            unique = sorted(set(scrapers))
            sample = ", ".join(unique[:6])
            extra = f" (+{len(unique) - 6} more)" if len(unique) > 6 else ""
            print(f"  • {check_name}: {len(scrapers)}× in {sample}{extra}")

    print("-" * 70 + "\n")


def print_summary_table(results: List[Dict]) -> None:
    """Print a human-readable table to stdout."""
    print("\n" + "=" * 70)
    print(f"{'SCRAPER':<30}  {'FILES':<6}  {'CHECKS':<10}  STATUS")
    print("=" * 70)
    for r in results:
        status  = PASS if r["all_passed"] else FAIL
        missing = MISS if r["files_found"] == 0 and not r.get("files_optional") else ""
        optional = " (optional)" if r.get("files_optional") and r["files_found"] == 0 else ""
        print(
            f"{r['scraper']:<30}  "
            f"{r['files_found']:<6}  "
            f"{r['checks_passed']}/{r['checks_total']:<7}  "
            f"{status}{optional} {missing}"
        )
    print("=" * 70)
    total_pass = sum(1 for r in results if r["all_passed"])
    print(f"\nTotal: {total_pass}/{len(results)} scrapers fully passed\n")


def write_github_summary(results: List[Dict], listing_date: str, report_date: str) -> None:
    """Write markdown to $GITHUB_STEP_SUMMARY if available."""
    summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if not summary_path:
        return
    lines = [
        f"## R2 Schema Monitor — listings {listing_date} (saved under {report_date})\n",
        "| Scraper | Files | Checks | Status |",
        "|---------|-------|--------|--------|",
    ]
    for r in results:
        icon   = PASS if r["all_passed"] else (MISS if r["files_found"] == 0 else FAIL)
        detail = ""
        if not r["all_passed"]:
            if r.get("files_optional") and r["files_found"] == 0:
                detail = "no files (optional)"
            elif r["files_found"] == 0:
                detail = "no Excel files found"
            else:
                failed_bits = []
                for f in r["file_results"]:
                    file_name = (f.get("file_key") or "?").split("/")[-1]
                    for c in f.get("validation", {}).get("checks", []):
                        if not c["passed"]:
                            bit = c["name"]
                            if c.get("detail"):
                                bit += f" ({c['detail']})"
                            failed_bits.append(f"{file_name}: {bit}")
                detail = "<br>".join(failed_bits[:8]) if failed_bits else "validation failed"
                if len(failed_bits) > 8:
                    detail += f"<br>… +{len(failed_bits) - 8} more"
        lines.append(
            f"| {r['scraper']} | {r['files_found']} | "
            f"{r['checks_passed']}/{r['checks_total']} | {icon} {detail} |"
        )
    total_pass = sum(1 for r in results if r["all_passed"])
    lines.append(f"\n**{total_pass}/{len(results)} scrapers fully passed**")
    with open(summary_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def write_github_alert_summary(alerts: List[Dict], run_date: str) -> None:
    """Append alert block to GitHub step summary when failures exist."""
    if not alerts:
        return
    summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if not summary_path:
        return
    lines = [
        f"\n### Alerts fired — {len(alerts)} issue(s)\n",
        "| Scraper | Severity | Check | Detail |",
        "|---------|----------|-------|--------|",
    ]
    for a in alerts[:40]:
        detail = a.get("detail", "").replace("|", "\\|")[:120]
        lines.append(
            f"| {a['scraper']} | {a['severity']} | {a['check']} | {detail} |"
        )
    if len(alerts) > 40:
        lines.append(f"\n_… and {len(alerts) - 40} more_")
    with open(summary_path, "a", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def upload_report(client, bucket: str, report: Dict, partition_date: str, site: Dict) -> None:
    """Upload the JSON report to R2 under the partition date (matches data folders)."""
    key  = report_r2_key(site, partition_date)
    body = json.dumps(report, ensure_ascii=False, indent=2, default=str).encode()
    try:
        client.put_object(Bucket=bucket, Key=key, Body=body, ContentType="application/json")
        log.info(f"Report uploaded → r2://{bucket}/{key}")
    except Exception as exc:
        log.warning(f"Could not upload report: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def parse_args():
    p = argparse.ArgumentParser(description="Inspect R2 Excel files against the project schema.")
    p.add_argument("--date",          default=None, help="YYYY-MM-DD to inspect (default: yesterday)")
    p.add_argument("--days-lookback", type=int, default=1, help="How many days to scan (default: 1 = yesterday only)")
    p.add_argument("--update-stats",  action="store_true", help="Write/update monitor_stats.yml with observed data")
    p.add_argument("--quality",       action="store_true", help="Run deep data-quality checks (slower)")
    p.add_argument("--verbose",     action="store_true", help="Log every failed check (no per-scraper cap)")
    p.add_argument("--fail-on-error", action="store_true", help="Exit with code 1 if any check fails")
    p.add_argument("--no-alert",      action="store_true", help="Skip webhook alerts even if URL is configured")
    p.add_argument(
        "--upload-config",
        action="store_true",
        help="Upload local websites-config.yml to R2 (bootstrap CI), then exit",
    )
    p.add_argument(
        "--upload-stats",
        metavar="PATH",
        default=None,
        help="Upload a local monitor_stats.yml to R2 (bootstrap trend baselines), then exit",
    )
    p.add_argument(
        "--site-slug",
        default=None,
        help="Override MONITOR_SITE_SLUG (e.g. 4sale) for this run",
    )
    return p.parse_args()


def main():
    args = parse_args()

    r2_client, bucket = build_r2_client()

    if args.upload_config or args.upload_stats:
        folder = resolve_site_folder(args.site_slug)
        site = load_site_config_from_r2(r2_client, bucket, folder)
        keys = monitor_data_keys(site)
        if args.upload_config:
            upload_config(r2_client, bucket, keys["config"])
            return
        if args.upload_stats:
            upload_stats_file(r2_client, bucket, Path(args.upload_stats), keys["stats"])
            return

    site = load_site_config_from_r2(r2_client, bucket, args.site_slug)
    keys = monitor_data_keys(site)

    # Date range to inspect
    if args.date:
        end_date   = datetime.strptime(args.date, "%Y-%m-%d")
    else:
        end_date   = datetime.utcnow() - timedelta(days=1)   # yesterday UTC
    dates_to_check = [end_date - timedelta(days=i) for i in range(args.days_lookback)]
    listing_date_str = end_date.strftime("%Y-%m-%d")
    report_date_str  = partition_date_for_data_date(end_date).strftime("%Y-%m-%d")

    log.info(
        f"Inspecting listing date(s): {[d.strftime('%Y-%m-%d') for d in dates_to_check]} "
        f"· report → monitor/{report_date_str}/report.json"
    )

    log.info(f"Connected to R2 bucket: {bucket}")
    log.info(f"Site folder: {site.get('folder')} · data prefix: {keys['base']}")

    config = load_config(r2_client, bucket, keys["config"])
    meta   = config.get("meta", {})
    alert_cfg = config.get("alert_thresholds", {})
    trend_cfg = alert_cfg.get("trend", {})

    # Build schema lookup: scraper_name → schema_entry
    schema_by_scraper = {
        e["scraper"]: e for e in config.get("excel_schema", [])
    }

    # Load historical stats from R2 for trend checks and/or --update-stats
    hist_stats = load_existing_stats(r2_client, bucket, keys["stats"])
    stats = dict(hist_stats) if args.update_stats else {}

    all_results  = []
    full_report  = {
        "run_date":     report_date_str,
        "inspect_date": listing_date_str,
        "folder":    site.get("folder"),
        "site_id":   site.get("site_id"),
        "website":   meta.get("website") or site.get("website"),
        "country":   meta.get("country") or site.get("country"),
        "repo":      meta.get("repo") or site.get("repo"),
        "display_name": site.get("display_name"),
        "meta":      meta,
        "scrapers":  {},
        "alerts":    [],
    }

    scrapers_cfg = config.get("scrapers", [])
    log.info(f"Processing {len(scrapers_cfg)} scrapers …")

    for scraper_cfg in scrapers_cfg:
        scraper_name = scraper_cfg["name"]
        r2_base      = r2_base_prefix(scraper_cfg.get("r2_path", ""))
        schema_entry = schema_by_scraper.get(scraper_name)

        if not r2_base:
            log.warning(f"  {scraper_name}: no r2_path in config — skipping")
            continue

        scraper_result = {
            "scraper":        scraper_name,
            "files_found":    0,
            "checks_passed":  0,
            "checks_total":   0,
            "all_passed":     True,
            "file_results":   [],
        }

        all_xlsx: List[Dict] = []
        seen_keys: set = set()
        tried_prefixes: List[str] = []
        for dt in dates_to_check:
            part_dt = partition_date_for_data_date(dt)
            for prefix in excel_prefixes_for_date(r2_base, part_dt):
                tried_prefixes.append(prefix)
                for f in list_excel_files(r2_client, bucket, prefix):
                    if f["key"] in seen_keys:
                        continue
                    seen_keys.add(f["key"])
                    f["date"] = dt.strftime("%Y-%m-%d")
                    all_xlsx.append(f)

        scraper_result["files_found"] = len(all_xlsx)

        if not all_xlsx:
            listing_date = dates_to_check[0].strftime("%Y-%m-%d")
            partition_date = partition_date_for_data_date(dates_to_check[0]).strftime("%Y-%m-%d")
            files_optional = (
                scraper_cfg.get("files_optional", False)
                or scraper_name in _FILES_OPTIONAL_SCRAPERS
            )
            sample = ", ".join(tried_prefixes[:2])
            extra  = f" (+{len(tried_prefixes) - 2} more)" if len(tried_prefixes) > 2 else ""
            if files_optional:
                log.info(
                    f"  {scraper_name}: no files for listing {listing_date} "
                    f"(optional — data is rare; partition day={partition_date})"
                )
                scraper_result["all_passed"] = True
                scraper_result["files_optional"] = True
            else:
                log.warning(
                    f"  {scraper_name}: NO Excel files found under {r2_base} "
                    f"for listing {listing_date} (R2 partition day={partition_date}); "
                    f"tried e.g. {sample}{extra}"
                )
                scraper_result["all_passed"] = False
            all_results.append(scraper_result)
            full_report["scrapers"][scraper_name] = scraper_result
            continue

        for xlsx_meta in all_xlsx:
            raw = download_excel(r2_client, bucket, xlsx_meta["key"])
            if raw is None:
                continue

            inspected = inspect_excel(raw, xlsx_meta["key"])
            inspected["size_bytes"] = xlsx_meta["size_bytes"]

            # Deep quality check on data sheets (if requested)
            if args.quality and inspected["readable"]:
                for sheet in inspected["sheets"]:
                    if sheet["name"] != "Info":
                        sheet["quality"] = check_data_quality(
                            raw, sheet["name"], listing_date_str
                        )

            # Validate against schema + alert floors + historical trends
            file_validation = {"file": xlsx_meta["key"], "checks": []}
            if schema_entry:
                schema_val = validate_file(inspected, schema_entry, scraper_name)
                floor_val = validate_observed_floor(inspected, scraper_name, alert_cfg)
                trend_val = validate_trends(
                    inspected,
                    scraper_name,
                    hist_stats.get(scraper_name),
                    trend_cfg,
                )
                file_validation = merge_validations(schema_val, floor_val, trend_val)
                file_validation["file"] = xlsx_meta["key"]
            else:
                log.debug(f"  {scraper_name}: no schema entry — skipping validation")

            scraper_result["file_results"].append(
                {**inspected, **{"validation": file_validation}}
            )
            scraper_result["checks_passed"] += sum(
                1 for c in file_validation.get("checks", []) if c["passed"]
            )
            scraper_result["checks_total"] += len(file_validation.get("checks", []))
            if not file_validation.get("passed", True):
                scraper_result["all_passed"] = False

            # Accumulate stats for --update-stats
            if args.update_stats:
                accumulate_stats(
                    stats,
                    scraper_name,
                    inspected,
                    xlsx_meta["size_bytes"],
                    xlsx_meta["date"],
                )

        all_results.append(scraper_result)
        full_report["scrapers"][scraper_name] = scraper_result
        status = PASS if scraper_result["all_passed"] else FAIL
        log.info(
            f"  {status} {scraper_name}: "
            f"{scraper_result['files_found']} file(s), "
            f"{scraper_result['checks_passed']}/{scraper_result['checks_total']} checks"
        )
        if not scraper_result["all_passed"]:
            log_scraper_failures(
                scraper_name,
                scraper_result,
                max_lines=None if args.verbose else _DEFAULT_MAX_FAILURE_LOGS,
            )

    # ── Outputs ───────────────────────────────────────────────────────────────
    alerts = collect_alerts(all_results, listing_date_str)
    full_report["alerts"] = alerts
    full_report["alert_count"] = len(alerts)

    print_summary_table(all_results)
    print_failure_summary(all_results)
    write_github_summary(all_results, listing_date_str, report_date_str)
    write_github_alert_summary(alerts, listing_date_str)
    upload_report(r2_client, bucket, full_report, report_date_str, site)

    if args.update_stats:
        save_stats(r2_client, bucket, stats, keys["stats"])

    # ── Webhook alerts ────────────────────────────────────────────────────────
    if not args.no_alert:
        webhook_url = os.environ.get("MONITOR_ALERT_WEBHOOK_URL", "").strip() or None
        try:
            dispatch_alerts(
                alerts, all_results, listing_date_str, alert_cfg, webhook_url, site, meta
            )
        except Exception as exc:
            log.error(f"Alert dispatch error (non-fatal): {exc}")

    # ── Exit code ─────────────────────────────────────────────────────────────
    any_failure = any(not r["all_passed"] for r in all_results)
    if any_failure and args.fail_on_error:
        log.error("One or more scrapers failed schema validation.")
        sys.exit(1)
    elif any_failure:
        log.warning("Some scrapers failed — exit 0 (use --fail-on-error to change).")


if __name__ == "__main__":
    main()
